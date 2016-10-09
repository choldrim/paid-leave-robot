import os
from datetime import datetime
from datetime import timedelta

import xlsxwriter
from openpyxl import load_workbook


class Tools:
    def get_excel_data(self, filename, columns=[], header_index=0):
        """
        header_index: the header row index default is 0 (first row)
        @return
            if the columns param is: ["name", "time"]
            {"name":[A, B, C...], "time": [8, 5, 10...]}
        """
        data = {}
        if not os.path.exists(filename):
            return data
        wb = load_workbook(filename=filename, read_only=True)
        s_name = wb.get_sheet_names()[0]
        ws = wb[s_name]

        header_cols = list(ws.rows)[header_index]  # get all headers
        for c in columns:
            for _header in header_cols:
                header_val = _header.value
                if header_val != c:
                    continue
                col_index = _header.column - 1
                vals = [list(row)[col_index].value for row in ws.rows]  # collect the specified col val from every row
                vals = list(vals)[header_index + 1:]  # remove the rows before header

                # for the openpyxl bug, return lots of None rows sometimes
                if len(vals) > 10000:
                    max_available = 0
                    for i in vals:
                        if i is None:
                            break
                        max_available += 1

                    data[c] = vals[0:max_available]
                else:
                    data[c] = vals

        return data

    def get_last_month_dt(self, cur_time):
        _t = cur_time.replace(day=1)
        last_month = _t - timedelta(days=1)
        last_month_dt = last_month.replace(day=1)
        return last_month_dt

    def get_month_str(self, t):
        return t.strftime("%Y%m")

    def write_to_execl(self, filename, cols=[[], ]):
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet()
        COLUMN = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L",
                  "M", "N", "O", "P", "Q", "R", "S", "T"]
        if len(cols[0]) > len(COLUMN):
            print("E: columns too long...")
            return False

        row = 1
        for col_data in cols:
            column = 0
            for value in col_data:
                pos = "%s%s" % (COLUMN[column], row)
                worksheet.write(pos, value)
                column += 1
            row += 1
        return True
